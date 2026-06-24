"""P&L агент Градекс KZ — ядро v2."""
import openpyxl
from collections import defaultdict

MONTH_COL='C'; THRESHOLD=300_000

def detect_div(name):
    n=(name or '').lower()
    if 'ауп' in n: return 'АУП'
    if 'кпо' in n: return 'КПО'
    if 'тшо' in n: return 'ТШО'
    if 'бетон' in n: return 'Бетон'
    if 'номенклатурная группа' in n: return 'Бетон'
    return None

def parse_1c(path):
    """Анализ счёта и полная ОСВ. P&L = счета 6xxx/7xxx. Подразделение распознаётся
    на любом уровне >=2, если есть активная статья (учитывает «плоские» субконто)."""
    wb=openpyxl.load_workbook(path); ws=wb.worksheets[0]
    def lvl(r):
        rd=ws.row_dimensions[r]; return rd.outline_level if rd else 0
    def is_pl(code):
        c=str(code).strip(); return c[:1] in ('6','7') and c[:4].isdigit()
    leaves=[]; acct=None; pl=False; item=None; item_val=None; item_kind=None; item_got=False; stack=[]
    def flush():
        if pl and item is not None and item_val and not item_got:
            leaves.append((acct,item,item_kind,None,float(item_val)))
    for r in range(1,ws.max_row+1):
        a=ws.cell(row=r,column=1).value
        if a is None: continue
        o=lvl(r)
        if a=='Итого' and o==0: continue
        e=ws.cell(row=r,column=5).value; f=ws.cell(row=r,column=6).value
        if o<=1:
            flush(); acct=str(a).split(',')[0].strip(); pl=is_pl(acct)
            item=None; item_val=None; item_got=False; stack=[]; continue
        if not pl: continue
        if a=='<...>':
            stack=[(lv,dv) for (lv,dv) in stack if lv<o]; continue
        div=detect_div(a)
        # строка-подразделение (на любом уровне) при активной статье -> разнесение
        if div and item is not None:
            stack=[(lv,dv) for (lv,dv) in stack if lv<o]
            anc=set(dv for (_,dv) in stack); val=e if e is not None else f
            if div not in anc and val:
                kind='income' if (f is not None and e is None) else 'expense'
                leaves.append((acct,item,kind,div,float(val))); item_got=True
            stack.append((o,div)); continue
        if o==2:   # новая статья (не подразделение)
            flush(); item=a; stack=[]
            item_val=e if e is not None else f
            item_kind='income' if (f is not None and e is None) else 'expense'
            item_got=False; continue
        stack=[(lv,dv) for (lv,dv) in stack if lv<o]; stack.append((o,None))
    flush()
    return leaves

def norm(s): return (s or '').strip().lower()

INCOME_MAP={
 ('6010',norm('1доход от реализации товаров')):('Бетон','INC_CONCRETE'),
 ('6010',norm('2Доход от сдачи в аренду')):('Бетон','INC_RENT'),
 ('6010',norm('Доход от реализации продукции и оказания услуг')):('ТШО','INC_SALES'),
 ('6110',norm('Вознаграждение по депозиту')):('ТШО','INC_FIN'),
 ('6230',norm('Доход от гос субсидий')):('Бетон','INC_SUBSIDY'),
}
OPEX_CONCEPT={
 'gps':'GPS','амортизация фа':'AMORT','аренда офиса':'RENT_OFFICE','аренда спецтехники':'RENT_TECH',
 'аренда транспорта':'RENT_TECH','база расходы по содержанию':'BASE_MAINT','гсм':'FUEL','зарплата':'SALARY',
 'интернет':'INTERNET','услуги связи':'INTERNET','канц товары':'OTHER','ком.услуги':'UTILITIES',
 'электроэнергия':'UTILITIES','медицинские услуги':'MEDICAL',
 'обязательные пенсионные взносы работодателя':'TAXES','отчисления осмс':'TAXES',
 'социальные отчисления':'TAXES','социальный налог':'TAXES','прочие расходы':'OTHER',
 'расходы на корреспонденцию':'DELIVERY','расходы на наем жилого помещения':'LODGING','суточные':'LODGING',
 'расходы на питание':'MEALS','расходы проживание':'MEALS','расходы на проезд':'TRAVEL',
 'расходы по доставке':'DELIVERY','расходы по обучению':'TRAINING','расходы по перевозке':'TRANSPORT',
 'расходы по сиз':'PPE','ремонт авто':'REPAIR','расходы по ремонту оборудования':'REPAIR',
 'ремонт оборудования':'REPAIR','расходы по ремонту':'REPAIR','расходы по ремонту авто':'REPAIR',
 'расходы по экологии':'UTIL','экология':'UTIL','расходы по утилизации':'UTIL','утилизация':'UTIL',
 'расходы на билеты':'TRAVEL','билеты':'TRAVEL','авиабилеты':'TRAVEL','аренда оборудования':'RENT_TECH','геодезические услуги':'SUBCONTRACT',
 'технические услуги':'SUBCONTRACT','расходы по технической диагностике':'SUBCONTRACT',
 'техническая диагностика':'SUBCONTRACT','тех диагностика':'SUBCONTRACT',
 'себестоимость реализованной продукции и оказанных услуг':'COGS','содержание офиса':'SITE_MAINT',
 'списание материалов':'MATERIALS','страхование':'INSURANCE',
}
OPEX_LABEL={
 'TCO':{'GPS':'GPS','AMORT':'Амортизация','RENT_OFFICE':'Аренда офиса','RENT_TECH':'Аренда техники','FUEL':'ГСМ',
   'SALARY':'Заработная плата','INTERNET':'Интернет+ связь','MEDICAL':'Мед.обслуживание','TAXES':'Налоги',
   'OTHER':'Прочие услуги сторонних организаций','LODGING':'Расходны на наем жилья и Суточные',
   'MEALS':'Расходы на питание и проживание','TRAVEL':'Расходы на проезд','DELIVERY':'Расходы по доставке',
   'TRAINING':'Расходы по обучению','TRANSPORT':'Расходы по перевозке','PPE':'Расходы по СИЗ','UTIL':'Расходы на утилизацию',
   'REPAIR':'Расходы на ремонт авто/ оборудования','COGS':'Услуги сторонних/ подрядных организаций',
   'SITE_MAINT':'Содержание на сайте офиса+ контейнера','MATERIALS':'Списанные материалы','INSURANCE':'Страхование',
   'UTILITIES':'Содержание на сайте офиса+ контейнера','BASE_MAINT':'Прочие услуги сторонних организаций','SUBCONTRACT':'Услуги сторонних/ подрядных организаций'},
 'KPO':{'GPS':'GPS','AMORT':'Амортизация','RENT_OFFICE':'Аренда офиса','RENT_TECH':'Аренда техники','FUEL':'ГСМ',
   'SALARY':'Заработная плата','INTERNET':'Интернет+связь','MEDICAL':'Мед.обслуживание','TAXES':'Налоги',
   'OTHER':'Прочие услуги','LODGING':'Расходны на наем жилья и Суточные','MEALS':'Расходы на питание и проживание',
   'TRAVEL':'Расходы на проезд','DELIVERY':'Расходы по доставке','TRAINING':'Расходы по обучению',
   'TRANSPORT':'Расходы по перевозке','PPE':'Расходы по СИЗ','UTIL':'Расходы на утилизацию',
   'REPAIR':'Расходы на ремонт авто/ оборудования','COGS':'Услуги сторонних/ подрядных организаций',
   'SITE_MAINT':'Содержание на сайте офиса+ контейнера','MATERIALS':'Списанные материалы','INSURANCE':'Страхование',
   'UTILITIES':'Содержание на сайте офиса+ контейнера','BASE_MAINT':'Прочие услуги','SUBCONTRACT':'Услуги сторонних/ подрядных организаций'},
 'CONCRETE':{'GPS':'GPS','RENT_TECH':'Аренда техники','BASE_MAINT':'База содержание','FUEL':'ГСМ',
   'SALARY':'Заработная плата','INTERNET':'Интернет+связь','UTILITIES':'Коммунальные услуги','TAXES':'Налоги',
   'OTHER':'Прочие расходы','LODGING':'Расходны на наем жилья (Эркан)  и Суточные','TRAVEL':'Расходы на проезд',
   'TRANSPORT':'Расходы по перевозке','PPE':'Расходы по СИЗ','UTIL':'Расходы по экологии',
   'REPAIR':'Расходы по ремонту авто/ оборудования','COGS':'Себестоимость товарного бетона',
   'MATERIALS':'Списанные материалы','INSURANCE':'Страхование','MEALS':'Прочие расходы','DELIVERY':'Прочие расходы',
   'TRAINING':'Прочие расходы','MEDICAL':'Прочие расходы','RENT_OFFICE':'Прочие расходы','SITE_MAINT':'Прочие расходы',
   'AMORT':'Амортизация','SUBCONTRACT':'Прочие расходы'},
}
INCOME_LABEL={
 'TCO':{'INC_SALES':'Доход от реализации','INC_FIN':'__OTHER_INCOME__','INC_OTHER':'__OTHER_INCOME__'},
 'KPO':{'INC_SALES':'Доход от реализации','INC_FIN':'__OTHER_INCOME__','INC_OTHER':'__OTHER_INCOME__'},
 'CONCRETE':{'INC_CONCRETE':'Доход от реализации бетона','INC_RENT':'Доход от сдачи в аренду',
   'INC_FIN':'Прочий доход','INC_SUBSIDY':'Прочие доходы (субсидии Даму)','INC_OTHER':'Прочий доход'},
}
ADMIN_CONCEPT={
 'зарплата':'Расходы по заработной плате','обязательные пенсионные взносы работодателя':'Налоги',
 'отчисления осмс':'Налоги','социальные отчисления':'Налоги','социальный налог':'Налоги','гсм':'ГСМ',
 'комиссия банка':'Банковская комиссия','амортизация фа':'Амортизация ФА','подписка':'Подписка',
 'интернет':'Услуги связи+интернет','услуги связи':'Услуги связи+интернет','медицинские услуги':'Мед осмотр',
 'страхование':'Расходы по страховке','суточные':'Командировочные расходы(проезд+проживание+суточные)',
 'расходы на наем жилого помещения':'Командировочные расходы(проезд+проживание+суточные)',
 'расходы на проезд':'Командировочные расходы(проезд+проживание+суточные)','содержание офиса':'Содержание офиса',
 'обучение сотрудников':'Обучение сотрудников','расходы по обучению':'Обучение сотрудников','расходы на билеты':'Командировочные расходы(проезд+проживание+суточные)','билеты':'Командировочные расходы(проезд+проживание+суточные)','канц товары':'Прочие расходы','прочие расходы':'Прочие расходы',
 'расходы на корреспонденцию':'Услуги по доставке корресп.и др.',
}
DIV_KEY={'ТШО':'TCO','КПО':'KPO','Бетон':'CONCRETE'}

def block_index(ws,start,end):
    idx={}
    for r in range(start,end+1):
        v=ws.cell(row=r,column=2).value
        if v is not None: idx[str(v).strip()]=r
    return idx

def other_income_row(ws):
    for r in range(28,40):
        v=ws.cell(row=r,column=2).value
        if v and 'прочие доходы' in str(v).lower(): return r
    return 31

def fill_report(template_path,leaves,out_path,month_col='C'):
    wb=openpyxl.load_workbook(template_path)
    tco,kpo,con,alls=wb['P&L 2026 TCO'],wb['P&L KPO'],wb['P&L 2026 concrete'],wb['P&L 2026 all']
    SH={'TCO':tco,'KPO':kpo,'CONCRETE':con}
    OPX={'TCO':block_index(tco,7,30),'KPO':block_index(kpo,7,30),'CONCRETE':block_index(con,9,28)}
    INC={'TCO':block_index(tco,3,5),'KPO':block_index(kpo,3,5),'CONCRETE':block_index(con,3,6)}
    OTH={'TCO':other_income_row(tco),'KPO':other_income_row(kpo)}
    ADM=block_index(con,31,46)
    BELOWPR={'TCO':block_index(tco,32,33)['Прочие расходы'],'KPO':block_index(kpo,32,33)['Прочие расходы'],'CONCRETE':block_index(con,48,49)['Прочие расходы']}
    acc=defaultdict(float); flags=[]; log=[]
    def add(ws,row,amt): acc[(ws.title,row)]+=amt
    def by_label(pool,label):
        if not label: return None
        lab=label.strip()
        for k,r in pool.items():
            if k.strip()==lab: return r
        return None
    def put(dk,label,amt,is_income=False):
        ws=SH[dk]
        if label=='__OTHER_INCOME__': add(ws,OTH[dk],amt); return ws.cell(row=OTH[dk],column=2).value
        if label=='__CONCRETE_ADMIN_AMORT__': add(con,ADM['Амортизация ФА'],amt); return 'Амортизация ФА (Бетон, адм.)'
        pool=INC[dk] if is_income else OPX[dk]
        r=by_label(pool,label)
        if r: add(ws,r,amt); return label.strip()
        return None

    for (account,item,kind,div,amount) in leaves:
        account=account or ''
        nit=norm(item)
        if kind=='income':
            d,concept=INCOME_MAP.get((account,nit),(div,'INC_OTHER'))
            if d not in DIV_KEY: d='Бетон'
            dk=DIV_KEY[d]; lab=put(dk,INCOME_LABEL[dk].get(concept,'__OTHER_INCOME__'),amount,True)
            log.append((account,item,div,'доход',SH[dk].title,lab,amount)); continue
        # КПН (77xx) считается формулой шаблона — пропускаем
        if account[:2]=='77':
            log.append((account,item,div or '—','кпн(пропуск)','—','—',amount)); continue
        # Финансирование 73xx -> операционная строка «Расходы на финансирование» по подразделению
        if account[:2]=='73':
            if div in DIV_KEY: dk=DIV_KEY[div]
            else: dk='CONCRETE'; flags.append((item,div or '—',amount,'финансирование без подразделения → concrete'))
            lab=put(dk,'Расходы на финансирование',amount)
            log.append((account,item,div or '—','расход',SH[dk].title,lab,amount)); continue
        # 7400 (курсовые 7430, обмен валюты 7480 и пр.) -> строка «Прочие расходы» (ниже операц.)
        if account[:2]=='74':
            if div in DIV_KEY: dk=DIV_KEY[div]
            else: dk='CONCRETE'; flags.append((item,div or '—',amount,'курсовые/прочие без подразделения → concrete'))
            add(SH[dk], BELOWPR[dk], amount)
            log.append((account,item,div or '—','прочие расх.',SH[dk].title,'Прочие расходы',amount)); continue
        if div=='АУП':
            label=ADMIN_CONCEPT.get(nit)
            if label and label in ADM:
                add(con,ADM[label],amount); log.append((account,item,div,'админ',con.title,label,amount))
            else:
                if amount>=THRESHOLD: flags.append((item,div,amount,'новая АУП-статья ≥300к'))
                add(con,ADM['Прочие расходы'],amount); log.append((account,item,div,'админ',con.title,'Прочие расходы',amount))
            continue
        if div not in DIV_KEY:
            flags.append((item, div or '—', amount, 'статья без подразделения → concrete, проверьте'))
            div='Бетон'
        dk=DIV_KEY[div]; concept=OPEX_CONCEPT.get(nit)
        if concept:
            lab=put(dk,OPEX_LABEL[dk].get(concept),amount)
            if lab is None:
                if amount>=THRESHOLD: flags.append((item,div,amount,f'≥300к, нет строки на {SH[dk].title}'))
                lab=put(dk,OPEX_LABEL[dk].get('OTHER'),amount)
            log.append((account,item,div,'расход',SH[dk].title,lab,amount))
        else:
            if amount>=THRESHOLD: flags.append((item,div,amount,f'новая статья ≥300к на {SH[dk].title}'))
            lab=put(dk,OPEX_LABEL[dk].get('OTHER'),amount)
            log.append((account,item,div,'расход(нов)',SH[dk].title,lab,amount))

    name2ws={ws.title:ws for ws in (tco,kpo,con)}
    for (sn,row),val in acc.items(): name2ws[sn][f'{month_col}{row}']=round(val,2)
    wire_all_sheet(alls,tco,kpo,con)
    wb.save(out_path)
    return acc,flags,log

def wire_all_sheet(alls,tco,kpo,con):
    def find(ws,text,a,b):
        for r in range(a,b+1):
            v=ws.cell(row=r,column=2).value
            if v and str(v).strip()==text: return r
        return None
    r5=find(alls,'Прочий доход',3,8); toi=other_income_row(tco); koi=other_income_row(kpo)
    if r5:
        for col in 'CDEFGHIJKLMN':
            alls[f'{col}{r5}']=f"='P&L 2026 concrete'!{col}5+'P&L 2026 TCO'!{col}{toi}+'P&L KPO'!{col}{koi}"
    a0=find(alls,'Расходы по заработной плате',35,52); c0=find(con,'Расходы по заработной плате',28,45)
    if a0 and c0:
        for i in range(16):
            for col in 'CDEFGHIJKLMN':
                alls[f'{col}{a0+i}']=f"='P&L 2026 concrete'!{col}{c0+i}"

if __name__=='__main__':
    leaves=parse_1c('/mnt/user-data/uploads/январь_2026.xlsx')
    acc,flags,log=fill_report('/mnt/user-data/uploads/GRD_2022-2026_1С_.xlsx',leaves,'/home/claude/GRD_январь_заполнен.xlsx')
    print('cells:',len(acc),'flags:',len(flags))
    for f in flags: print('  FLAG:',f)
